"""
导出工具 - 支持导出成绩为 Excel/PDF
"""
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

class ExportUtils:
    @staticmethod
    def export_to_excel(data, filename, headers=None):
        """
        导出数据到 Excel
        data: list of dict
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "成绩单"
            
            # 如果没有提供表头，从第一行数据获取
            if not headers and data:
                headers = list(data[0].keys())
            
            # 写入表头
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            logger.info(f"Data exported to Excel: {filename}")
            return True, f"导出成功: {filename}"
            
        except ImportError:
            logger.error("openpyxl not installed")
            return False, "缺少 openpyxl 库，请安装: pip install openpyxl"
        except Exception as e:
            logger.error(f"Export to Excel failed: {e}")
            return False, f"导出失败: {str(e)}"

    @staticmethod
    def export_to_pdf(data, filename, title="成绩报告"):
        """
        导出数据到 PDF
        """
        try:
            from fpdf import FPDF
            
            pdf = FPDF()
            pdf.add_page()
            
            # 添加中文字体支持（需要字体文件）
            # pdf.add_font('SimSun', '', 'simsun.ttf', uni=True)
            # pdf.set_font('SimSun', '', 12)
            
            # 使用默认字体
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 10, title, 0, 1, 'C')
            pdf.ln(10)
            
            pdf.set_font('Arial', '', 10)
            
            # 写入数据
            for row in data:
                for key, value in row.items():
                    pdf.cell(0, 10, f"{key}: {value}", 0, 1)
                pdf.ln(5)
            
            pdf.output(filename)
            logger.info(f"Data exported to PDF: {filename}")
            return True, f"导出成功: {filename}"
            
        except ImportError:
            logger.error("fpdf not installed")
            return False, "缺少 fpdf 库，请安装: pip install fpdf"
        except Exception as e:
            logger.error(f"Export to PDF failed: {e}")
            return False, f"导出失败: {str(e)}"

    @staticmethod
    def generate_filename(prefix, extension):
        """生成带时间戳的文件名"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
